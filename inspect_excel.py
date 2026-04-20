"""
Inspect Excel structure to understand layout.
"""

import openpyxl
from pathlib import Path

test_file = Path(__file__).parent / "excel_files" / "15.2.1 - MP-FKR018-RIL-001 - PROJETO DE PISCINA E AQUECIMENTO.xlsx"

wb = openpyxl.load_workbook(test_file, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print("\nFirst 25 rows:")
print("="*200)

for row_idx in range(1, 26):
    row_values = []
    for col_idx in range(1, 20):
        cell = ws.cell(row_idx, col_idx)
        val = cell.value
        if val is None:
            row_values.append("")
        else:
            row_values.append(str(val)[:20])  # Truncate to 20 chars
    
    print(f"Row {row_idx:2d}: {' | '.join(row_values)}")

print("\n" + "="*200)
print("Looking for data table...")

for row_idx in range(1, ws.max_row + 1):
    row_values = []
    for col_idx in range(1, 5):
        cell = ws.cell(row_idx, col_idx)
        val = cell.value
        row_values.append(str(val) if val else "")
    
    row_text = " ".join(row_values).lower()
    if "item" in row_text and "descrição" in row_text:
        print(f"\n✅ Header row found at row {row_idx}")
        print(f"   Content: {[ws.cell(row_idx, c).value for c in range(1, 15)]}")
        
        # Show next 10 rows
        print("\nData rows:")
        for data_row in range(row_idx + 1, min(row_idx + 11, ws.max_row + 1)):
            row_data = [ws.cell(data_row, c).value for c in range(1, 15)]
            print(f"   Row {data_row}: {row_data}")
        break

wb.close()

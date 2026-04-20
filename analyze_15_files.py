import openpyxl
import os

# Get all 15.2 files
files = sorted([f for f in os.listdir('.') if f.startswith('15.2')])

# Check each file for unique sheets and structure
print('=== Excel Files and Their Sheets ===\n')
sheet_patterns = {}
for fname in files:
    try:
        wb = openpyxl.load_workbook(fname, data_only=True)
        sheets = ','.join(wb.sheetnames)
        sheet_patterns[fname] = sheets
        wb.close()
        print(f'{fname[:60]:60} => {len(wb.sheetnames)} sheet(s): {sheets}')
    except Exception as e:
        print(f'{fname[:60]:60} => ERROR: {str(e)[:30]}')

# Now analyze the first file in detail - all rows and columns
print('\n\n=== Detailed Analysis of First File ===')
fname = files[0]
wb = openpyxl.load_workbook(fname, data_only=True)
ws = wb.active

# Print all rows with data
print(f'\nAll data in sheet "{ws.title}":')
for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    # Skip empty rows
    if any(cell for cell in row):
        # Print with proper formatting
        formatted_row = [str(cell)[:20] if cell else '---' for cell in row]
        print('Row {:2}: {}'.format(row_idx, ' | '.join(formatted_row)))

wb.close()

# Now let's see if any file has multiple tables/sheets
print('\n\n=== Summary: Files with Multiple Sheets ===')
multi_sheet_files = [f for f, s in sheet_patterns.items() if ',' in s]
if multi_sheet_files:
    for fname in multi_sheet_files:
        print(f'  - {fname}')
else:
    print('  (None found - all files have single sheet)')

# Summary of structure
print(f'\n\nTotal 15.2 files: {len(files)}')
print(f'Files with multiple sheets: {len(multi_sheet_files)}')
print(f'\nPattern: This appears to be a "Mapa de Concorrência" budget/quote document')
print(f'(similar to existing orcamento template)')

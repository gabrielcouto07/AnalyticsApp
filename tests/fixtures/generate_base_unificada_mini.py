"""Gera tests/fixtures/base_unificada_mini.xlsx.

Réplica sintética (100% fictícia — nenhum dado real de cliente) do workbook com
uma aba 'Base Unificada' VÁLIDA, reproduzindo os "defeitos" estruturais do
arquivo real que o parser precisa tratar:

- cabeçalho NÃO na 1ª linha (2 linhas de resumo/título antes — como no real);
- coluna 'Cliente / Fornecedor' com espaços ao redor da barra (vs canônico);
- valores monetários já legíveis (ponto decimal) guardados como texto;
- Mês/Ano como texto MM/AAAA;
- CNPJ com zero à esquerda + 1 CNPJ intercompany;
- uma linha de cabeçalho REPETIDA no meio dos dados (região deve descartá-la);
- abas brutas (Saída/Entrada/Venda) presentes — o fato NÃO pode concatená-las
  com a Base (senão duplicaria registros).

Totais escolhidos à mão (verificados nos testes) — Base Unificada:
- Saída 01/2026 = 16 × 1000 = 16000.00 (uma delas intercompany)
- Saída 06/2025 = 5 × 500 = 2500.00
- Entrada 01/2026 = 3 × 100 = 300.00
- linhas de dados = 24
"""
from datetime import datetime
from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "base_unificada_mini.xlsx"

BU_HEADER = [
    "Mês/Ano", "Tipo Movimento", "Data do Documento", "Série", "Serial",
    "Cliente / Fornecedor", "CNPJ", "UF", "Município", "Grupo Item",
    "Descrição do Item", "Utilização", "Quantidade", "Valor (R$)", "Vendedor",
    "Ano", "Mês", "Linha de Negócio", "CNPJ Excluído",
]
INTERCOMPANY = "33921755000188"


def bu_row(mesano, tipo, data, serial, nome, cnpj, uf, grupo, qtd, valor, vend, ldn, excl):
    ano = int(mesano.split("/")[1]); mes = int(mesano.split("/")[0])
    return [mesano, tipo, data, "26", serial, nome, cnpj, uf, "Cidade",
            grupo, "ITEM DESC", "VENDA", qtd, valor, vend, ano, mes, ldn, excl]


def build_base_rows():
    rows = []
    # Saída 01/2026 — 16 linhas de 1000 (a última intercompany)
    for i in range(15):
        rows.append(bu_row("01/2026", "Saída", datetime(2026, 1, 5), str(100 + i),
                            f"CLIENTE {i}", "01234567000199", "SP", "EQUIPAMENTOS MORITA",
                            1, 1000.0, "VENDEDOR A", "MORITA", "Não"))
    rows.append(bu_row("01/2026", "Saída", datetime(2026, 1, 6), "199", "FILIAL INTERCO",
                        INTERCOMPANY, "MG", "PECAS CARESTREAM", 1, 1000.0, "VENDEDOR B",
                        "CARESTREAM", "Sim"))
    # Saída 06/2025 — 5 linhas de 500
    for i in range(5):
        rows.append(bu_row("06/2025", "Saída", datetime(2025, 6, 10), str(200 + i),
                            f"CLIENTE ANT {i}", "01234567000199", "SP", "EQUIPAMENTOS MORITA",
                            1, 500.0, "VENDEDOR A", "MORITA", "Não"))
    # Entrada 01/2026 — 3 linhas de 100
    for i in range(3):
        rows.append(bu_row("01/2026", "Entrada", datetime(2026, 1, 8), str(300 + i),
                            f"CLIENTE DEV {i}", "01234567000199", "SP", "EQUIPAMENTOS MORITA",
                            1, 100.0, "VENDEDOR A", "MORITA", "Não"))
    return rows


# Cabeçalho fiscal bruto reduzido (para as abas Saída/Entrada existirem e serem
# classificadas — mas NÃO devem ser somadas ao fato quando a Base é válida)
FISCAL_HEADER = ["Mês/Ano", "Nº Documento", "Série", "Serial", "Nome PN", "CNPJ",
                 "UF", "Municipio", "Grupo Item", "Descrição do item", "Data do doc.",
                 "Quantidade", "Valor contábil", "Valor mercadorias", "Gratuito"]


def fiscal_row(doc, valor):
    return ["01/2026", doc, "26", str(doc), "PN", "01234567000199", "SP", "Cidade",
            "EQUIPAMENTOS MORITA", "ITEM", datetime(2026, 1, 5), 1, valor, valor, False]


VENDA_HEADER = ["Mês/Ano", "Nº do documento", "Nome do PN", "Total do Documento",
                "Nome do vendedor", "Estado", "Utilização"]


def build():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["TOTAL SAÍDA", 999999])
    ws.append(["TOTAL ENTRADA", 111])

    ws = wb.create_sheet("Painel Gráficos")
    ws.append(["(gráficos prontos — ignorar)"])

    # Base Unificada VÁLIDA — cabeçalho na 3ª linha (índice 2)
    ws = wb.create_sheet("Base Unificada")
    ws.append(["Total linhas Saída:", 21, None, "Total linhas Entrada:", 3])  # linha resumo
    ws.append([None] * len(BU_HEADER))                                        # linha em branco
    ws.append(BU_HEADER)                                                      # cabeçalho real (linha 2)
    rows = build_base_rows()
    for i, r in enumerate(rows):
        ws.append(r)
        if i == 9:
            ws.append(BU_HEADER)   # cabeçalho REPETIDO no meio (região deve descartar)

    # Abas brutas — presentes, mas o fato deve vir só da Base (sem duplicar)
    ws = wb.create_sheet("Dados Saída")
    ws.append(FISCAL_HEADER)
    for d in range(5):
        ws.append(fiscal_row(1000 + d, 7777.0))   # totais propositalmente diferentes

    ws = wb.create_sheet("Dados Entrada")
    ws.append(FISCAL_HEADER)
    ws.append(fiscal_row(2000, 55.0))

    ws = wb.create_sheet("Dados Venda")
    ws.append(VENDA_HEADER)
    ws.append(["01/2026", 1000, "PN", 333.0, "VENDEDOR A", "SP", "VENDA"])

    ws = wb.create_sheet("Dados Linha de Negócio")
    ws.append(["Grupo Item", "Linha de Negócio"])
    ws.append(["EQUIPAMENTOS MORITA", "MORITA"])
    ws.append(["PECAS CARESTREAM", "CARESTREAM"])

    ws = wb.create_sheet("Leia-me")
    ws.append(["Notas do workbook — ignorar"])

    wb.save(OUT)
    print(f"fixture salvo em {OUT}")


if __name__ == "__main__":
    build()

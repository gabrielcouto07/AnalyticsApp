"""Gera o fixture tests/fixtures/medical_mini.xlsx.

Réplica em miniatura (e 100% sintética — nenhum dado real de cliente) do
workbook DASHBOARD_MEDICAL: mesmas abas, mesmos nomes de colunas-chave e os
mesmos "defeitos" que o parser precisa tratar (CPF numérico sem zero à
esquerda, Série numérica, Mês/Ano corrompido para data, aba Entrada com
linhas em branco, Venda sem Mês/Ano preenchido, etc.).

Totais escolhidos à mão (verificados nos testes):
- Saída 01/2026: 1000.00 + 2000.50 + 300.25 + 150.00 + 549.25 = 4000.00
- Saída 06/2025: 500.00 + 250.00 + 250.00 = 1000.00
- Entrada 01/2026 (devoluções): 100.00 + 150.00 = 250.00
- Venda: 1234.56 + 765.44 + 500.00 = 2500.00

Rodar: python tests/fixtures/generate_medical_mini.py
"""
from datetime import datetime
from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "medical_mini.xlsx"

# Cabeçalho fiscal reduzido: colunas usadas pela tabela fato + amostra de
# colunas de impostos zeradas (para exercitar o filtro de colunas úteis)
FISCAL_HEADER = [
    "Mês/Ano", "Tipo do objeto", "Processo", "Nº interno", "Nº Documento",
    "Série", "Serial", "Razão Social Entidade", "CNPJ Entidade",
    "Data do doc.", "Data de Criação", "Data do lanc.", "Data venc.",
    "Cód. PN", "Nome PN", "CNPJ", "CPF", "Inscr. Estadual (I.E)",
    "UF", "Cod. IBGE", "Municipio", "Código do item", "Descrição do item",
    "Grupo Item", "NCM", "CFOP", "Quantidade", "Utilização",
    "Valor mercadorias", "Valor contábil", "Gratuito",
    "Valor base ICMS", "% Alíq. ICMS", "Valor ICMS", "Valor base IPI",
    "Valor IPI", "Valor base PIS", "Valor PIS", "Valor base COFINS",
    "Valor COFINS", "Valor frete", "Valor seguro", "Valor desconto",
]

D = datetime  # atalho


def fiscal_row(mes_ano, doc, serie, serial, nome, cnpj, cpf, uf, municipio,
               item_desc, grupo, qtd, valor, data, processo="Nota Fiscal de Saída"):
    return [
        mes_ano, 13, processo, doc + 50000, doc,
        serie, serial, "EMPRESA TESTE LTDA", "09.020.873/0001-30",
        data, data, data, data,
        f"CLI{doc:06d}", nome, cnpj, cpf, "ISENTO",
        uf, "3106200", municipio, f"ITEM{doc:04d}", item_desc,
        grupo, "90183199", "6108", qtd, "VENDA CONTRIB",
        valor, valor, False,
        0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
    ]


SAIDA_ROWS = [
    # 01/2026 — soma Valor contábil = 4000.00
    fiscal_row("01/2026", 100, 26, "75", "CLINICA ALFA LTDA", "11111111000191", None,
               "MG", "Belo Horizonte", "EQUIPAMENTO RAIO-X A", "EQUIPAMENTOS MORITA", 1, 1000.00, D(2026, 1, 5)),
    # Mês/Ano corrompido para data pelo Excel — o parser deve recuperar "01/2026"
    fiscal_row(D(2026, 1, 1), 101, 26, "76", "CLINICA BETA LTDA", "22222222000108", None,
               "SP", "Campinas", "SENSOR DIGITAL B", "PECAS CARESTREAM", 2, 2000.50, D(2026, 1, 10)),
    # CPF numérico com zero à esquerda perdido (deve virar '08076748852')
    fiscal_row("01/2026", 102, 26, "77", "DR EXEMPLO PESSOA FISICA", None, 8076748852,
               "RJ", "Rio de Janeiro", "FILME PERIAPICAL", "FILME DENTAL", 10, 300.25, D(2026, 1, 15)),
    # CNPJ intercompany — deve marcar 'CNPJ Excluído' = 'Sim'
    fiscal_row("01/2026", 103, 26, "78", "FILIAL INTERCOMPANY SA", "33921755000188", None,
               "MG", "Contagem", "TRANSFERENCIA INTERNA", "PECAS CARESTREAM", 1, 150.00, D(2026, 1, 20)),
    # Grupo Item ausente do de-para — deve virar 'NÃO MAPEADO'
    fiscal_row("01/2026", 104, 26, "79", "HOSPITAL GAMA", "44444444000109", None,
               "BA", "Salvador", "PRODUTO SEM GRUPO", "GRUPO NOVO XYZ", 3, 549.25, D(2026, 1, 25)),
    # 06/2025 — soma = 1000.00 (para o comparativo anual)
    fiscal_row("06/2025", 105, 26, "80", "CLINICA ALFA LTDA", "11111111000191", None,
               "MG", "Belo Horizonte", "EQUIPAMENTO RAIO-X A", "EQUIPAMENTOS MORITA", 1, 500.00, D(2025, 6, 5)),
    fiscal_row("06/2025", 106, 26, "81", "CLINICA DELTA", "55555555000105", None,
               "PR", "Curitiba", "SENSOR DIGITAL B", "PECAS CARESTREAM", 1, 250.00, D(2025, 6, 12)),
    fiscal_row("06/2025", 107, 26, "82", "CLINICA BETA LTDA", "22222222000108", None,
               "SP", "Campinas", "FILME PERIAPICAL", "FILME DENTAL", 5, 250.00, D(2025, 6, 20)),
]

ENTRADA_ROWS = [
    fiscal_row("01/2026", 200, 1, "10", "CLINICA ALFA LTDA", "11111111000191", None,
               "MG", "Belo Horizonte", "EQUIPAMENTO RAIO-X A", "EQUIPAMENTOS MORITA", 1, 100.00,
               D(2026, 1, 8), processo="Nota Fiscal de Saída ( Devolução )"),
    fiscal_row("01/2026", 201, 1, "11", "HOSPITAL GAMA", "44444444000109", None,
               "BA", "Salvador", "PRODUTO SEM GRUPO", "GRUPO NOVO XYZ", 1, 150.00,
               D(2026, 1, 22), processo="Nota Fiscal de Saída ( Devolução )"),
]

VENDA_HEADER = [
    "Mês/Ano", "Documento", "Nome da filial", "Código do PN", "Nome do PN",
    "Nº NF", "Nº do documento", "Nº interno", "Utilização", "Nome do vendedor",
    "Estado", "Qtde de NF's/Dev NF's", "Adiantamento", "Total a pagar",
    "Total do Documento", "Total pago", "Total com IRF (sem adiantame",
] + [f"Característica {i}" for i in range(1, 11)]

VENDA_ROWS = [
    # Mês/Ano vazio — deve ser derivado da Saída (doc 100 → 01/2026)
    [None, "NF", "EMPRESA TESTE", "CLI000100", "CLINICA ALFA LTDA", 900, 100, 50100,
     "VENDA CONTRIB", "VENDEDOR A", "MG", 1, 0, 0, 1234.56, 1234.56, 1234.56] + [None] * 10,
    # Mês/Ano preenchido na própria aba (tem prioridade)
    ["06/2025", "NF", "EMPRESA TESTE", "CLI000105", "CLINICA ALFA LTDA", 901, 105, 50105,
     "VENDA CONTRIB", "VENDEDOR B", "MG", 1, 0, 0, 765.44, 765.44, 765.44] + [None] * 10,
    # Documento sem correspondência na Saída — Mês/Ano permanece vazio
    [None, "Dev NF", "EMPRESA TESTE", "CLI009999", "CLIENTE SEM NOTA", 902, 9999, 59999,
     "SAID BONIFIC BRINDE", "VENDEDOR A", "SP", 1, 0, 0, 500.00, 500.00, 500.00] + [None] * 10,
]

LOOKUP_ROWS = [
    ("Grupo Item", "Linha de Negócio"),
    ("EQUIPAMENTOS MORITA", "MORITA"),
    ("PECAS CARESTREAM", "CARESTREAM"),
    ("FILME DENTAL", "CARESTREAM"),
    ("SERVICOS GENORAY", "GENORAY"),
]


def build() -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Painel Gráficos"
    ws.append(["(gráficos prontos — o app deve ignorar esta aba)"])

    ws = wb.create_sheet("Dashboard")
    ws.append(["TOTAL SAÍDA", 5000])
    ws.append(["TOTAL ENTRADA", 250])

    ws = wb.create_sheet("Dados Saída")
    ws.append(FISCAL_HEADER)
    for row in SAIDA_ROWS:
        ws.append(row)

    ws = wb.create_sheet("Dados Entrada")
    ws.append(FISCAL_HEADER)
    for row in ENTRADA_ROWS:
        ws.append(row)
    # linhas fantasma em branco (a aba real tem ~1.100 delas)
    for _ in range(5):
        ws.append([None] * len(FISCAL_HEADER))

    ws = wb.create_sheet("Dados Venda")
    ws.append(VENDA_HEADER)
    for row in VENDA_ROWS:
        ws.append(row)

    ws = wb.create_sheet("Dados Linha de Negócio")
    for row in LOOKUP_ROWS:
        ws.append(row)

    ws = wb.create_sheet("Base Unificada")
    ws.append(["Total linhas Saída:", len(SAIDA_ROWS)])
    ws.append(["(staging manual — o app deve ignorar esta aba)"])

    ws = wb.create_sheet("Leia-me")
    ws.append(["Notas do workbook — o app deve ignorar esta aba"])

    wb.save(OUT)
    print(f"fixture salvo em {OUT}")


if __name__ == "__main__":
    build()

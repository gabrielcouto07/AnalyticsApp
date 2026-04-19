"""
Orcamento Parser — Extracts structured data from "Mapa de Concorrência" Excel files.

Layout (consistent across all files):
  Row 2:  "MAPA DE CONCORRÊNCIA" title (col A or B)
  Row 6:  "FORNECEDOR 1" at col 10, "FORNECEDOR 2" at col 13, "FORNECEDOR 3" at col 16...
          Each fornecedor block spans 3 columns: [label_col, value_col, gap]
          → label_col = start_col (NOME, CONTATO, TELEFONE, EMAIL labels)
          → value_col = start_col + 1 (actual values)
  Row 8:  OBRA (C3), N.º (C8), NOME fornecedor (start_col+1 per block)
  Row 9:  DATA (C8), CONTATO (start_col+1)
  Row 10: ASSU. (C3), TELEFONE (start_col+1)
  Row 11: EMAIL (start_col+1)
  Row 13: Item headers: ITEM(C2), DESCRIÇÃO(C3), QUANT(C4), UNID(C5), ...
          + "PREÇOS" at each fornecedor start_col
  Row 14: Price sub-headers per fornecedor:
          start_col = price_col_a (e.g. "VALOR UNIT." or "VALOR INICIAL")
          start_col+1 = price_col_b (e.g. "VALOR TOTAL" or "VALOR NEGOCIADO")
          → price_col_b is ALWAYS the "final/important" price
  Row 17+: Item data rows until total/empty
  Total row: has "TOTAL" or "SALDO" in col 8 or col 10

Year/date from cell C8 row 9 (datetime).
Assunto from cell C3 row 10.
Obra from cell C3 row 8.
"""

import re
import pandas as pd
import numpy as np
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import openpyxl


# ─── Fornecedor column positions ──────────────────────────────────────────────
# FORNECEDOR 1 starts at col 10, each block is 3 cols apart
FORNECEDOR_START_COL = 10
FORNECEDOR_STRIDE = 3
MAX_FORNECEDORES = 5  # scan up to 5


def _parse_num(v) -> Optional[float]:
    """Parse a cell value to float. Returns None if not numeric."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if np.isnan(v) or np.isinf(v):
            return None
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "_", "NA", "nan", "None", "#REF!", "#VALUE!", "SALDO"):
        return None
    # Remove R$, spaces
    s = re.sub(r"[R$\s]", "", s)
    # BR format: 1.234,56 → 1234.56
    if re.search(r"\d\.\d{3},\d", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _safe_str(v) -> str:
    """Safely convert cell value to stripped string."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


def _find_fornecedor_blocks(ws) -> List[Dict[str, Any]]:
    """
    Scan row 6 to find all FORNECEDOR headers and their column positions.
    Returns list of {index: int, start_col: int, label: str}
    """
    blocks = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=6, column=col).value
        if v and "FORNECEDOR" in str(v).upper():
            blocks.append({
                "index": len(blocks) + 1,
                "start_col": col,
                "label": _safe_str(v),
            })
    return blocks


def _extract_fornecedor_info(ws, start_col: int) -> Dict[str, str]:
    """Extract NOME, CONTATO, TELEFONE, EMAIL for one fornecedor block."""
    val_col = start_col + 1
    return {
        "Nome": _safe_str(ws.cell(row=8, column=val_col).value),
        "Contato": _safe_str(ws.cell(row=9, column=val_col).value),
        "Telefone": _safe_str(ws.cell(row=10, column=val_col).value),
        "Email": _safe_str(ws.cell(row=11, column=val_col).value),
    }


def _extract_price_headers(ws, start_col: int) -> Tuple[str, str]:
    """Extract the two price sub-header labels from row 14."""
    h1 = _safe_str(ws.cell(row=14, column=start_col).value)
    h2 = _safe_str(ws.cell(row=14, column=start_col + 1).value)
    return h1 or "Valor A", h2 or "Valor B"


def _extract_meta(ws) -> Dict[str, Any]:
    """Extract file-level metadata: Obra, Assunto, Numero, Data."""
    obra = _safe_str(ws.cell(row=8, column=3).value)
    assunto = _safe_str(ws.cell(row=10, column=3).value)
    numero = _safe_str(ws.cell(row=8, column=8).value)

    data_val = ws.cell(row=9, column=8).value
    data_str = ""
    if data_val:
        try:
            if hasattr(data_val, "strftime"):
                data_str = data_val.strftime("%Y-%m-%d")
            else:
                data_str = str(data_val)
        except Exception:
            data_str = str(data_val)

    return {
        "Obra": obra,
        "Assunto": assunto,
        "Numero": numero,
        "Data": data_str,
    }


def _is_item_row(ws, row: int) -> bool:
    """Check if a row is an item row (has numeric ITEM in col 2 and description in col 3)."""
    item_val = ws.cell(row=row, column=2).value
    desc_val = ws.cell(row=row, column=3).value
    if item_val is None and desc_val is None:
        return False
    # Item number should be numeric
    if item_val is not None:
        try:
            int(float(str(item_val)))
            return True
        except (ValueError, TypeError):
            pass
    # Sometimes item number is missing but description exists with prices
    if desc_val and _safe_str(desc_val) not in ("", "ENDEREÇO OBRA:"):
        # Check if any price column has a value
        for col_offset in range(FORNECEDOR_START_COL, FORNECEDOR_START_COL + MAX_FORNECEDORES * FORNECEDOR_STRIDE, FORNECEDOR_STRIDE):
            for sub in (0, 1):
                pv = _parse_num(ws.cell(row=row, column=col_offset + sub).value)
                if pv is not None:
                    return True
    return False


def _is_total_row(ws, row: int) -> bool:
    """Check if row is a total/saldo row.
    Only col 8 (SALDO TOTAL) is reliable — fornecedor columns (10, 13, 16)
    sometimes use 'SALDO' as a price label, not as a total indicator.
    A total row in col 8 typically has 'TOTAL' text or a numeric sum,
    AND has no item number in col 2.
    """
    item_val = ws.cell(row=row, column=2).value
    # If col 2 has a valid item number, this is NOT a total row
    if item_val is not None:
        try:
            int(float(str(item_val)))
            return False  # valid item number → not a total row
        except (ValueError, TypeError):
            pass

    v8 = _safe_str(ws.cell(row=row, column=8).value).upper()
    if v8 in ("TOTAL", "SALDO", "SALDO COM DESCONTO"):
        return True

    # Also check if col 8 has a numeric value (total sum) with no item in col 2
    if item_val is None and _parse_num(ws.cell(row=row, column=8).value) is not None:
        desc = _safe_str(ws.cell(row=row, column=3).value)
        if not desc or desc.upper().startswith("ENDEREÇO"):
            return True

    return False


def parse_orcamento_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    Parse a Mapa de Concorrência Excel file.

    Returns dict with:
        meta: {Obra, Assunto, Numero, Data, Filename}
        orcamento: [{Item, Descricao, Quant, Unid, Tipo}, ...]
        fornecedores: [{Nome, Contato, Telefone, Email, PriceHeaderA, PriceHeaderB}, ...]
        price_map: [{Item, FornecedorIndex, FornecedorNome, ValorA, ValorB}, ...]
        flat: DataFrame with one row per (Item × Fornecedor) — ready for analytics
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # ── Meta ──────────────────────────────────────────────────────────
    meta = _extract_meta(ws)
    meta["Filename"] = filename

    # ── Fornecedores ──────────────────────────────────────────────────
    forn_blocks = _find_fornecedor_blocks(ws)
    fornecedores = []
    for block in forn_blocks:
        info = _extract_fornecedor_info(ws, block["start_col"])
        price_h = _extract_price_headers(ws, block["start_col"])
        fornecedores.append({
            "FornecedorIndex": block["index"],
            "Nome": info["Nome"],
            "Contato": info["Contato"],
            "Telefone": info["Telefone"],
            "Email": info["Email"],
            "PriceHeaderA": price_h[0],
            "PriceHeaderB": price_h[1],
            "StartCol": block["start_col"],
        })

    # ── Orçamento Items + Price Map ───────────────────────────────────
    orcamento = []
    price_map = []

    # Items start at row 17, go until total or end
    for row in range(17, ws.max_row + 1):
        if _is_total_row(ws, row):
            break

        if not _is_item_row(ws, row):
            continue

        item_num = _parse_num(ws.cell(row=row, column=2).value)
        item_id = int(item_num) if item_num is not None else len(orcamento) + 1
        descricao = _safe_str(ws.cell(row=row, column=3).value)
        quant = _parse_num(ws.cell(row=row, column=4).value) or 0
        unid = _safe_str(ws.cell(row=row, column=5).value)

        # Type: Serviço if quant == 1, else Insumo
        tipo = "Serviço" if quant == 1 else "Insumo"

        orcamento.append({
            "Item": item_id,
            "Descricao": descricao,
            "Quant": quant,
            "Unid": unid,
            "Tipo": tipo,
        })

        # Extract prices from each fornecedor
        for forn in fornecedores:
            sc = forn["StartCol"]
            val_a = _parse_num(ws.cell(row=row, column=sc).value)
            val_b = _parse_num(ws.cell(row=row, column=sc + 1).value)

            price_map.append({
                "Item": item_id,
                "Descricao": descricao,
                "FornecedorIndex": forn["FornecedorIndex"],
                "FornecedorNome": forn["Nome"],
                "ValorA": val_a,
                "ValorB": val_b,
                "Preco": val_b if val_b is not None else val_a,
            })

    # ── Build flat DataFrame ──────────────────────────────────────────
    flat_records = []
    for orc_item in orcamento:
        for forn in fornecedores:
            # Find matching price
            matching = [
                p for p in price_map
                if p["Item"] == orc_item["Item"]
                and p["FornecedorIndex"] == forn["FornecedorIndex"]
            ]
            preco = matching[0]["Preco"] if matching else None
            val_a = matching[0]["ValorA"] if matching else None
            val_b = matching[0]["ValorB"] if matching else None

            flat_records.append({
                # Meta
                "Obra": meta["Obra"],
                "Assunto": meta["Assunto"],
                "Numero": meta["Numero"],
                "Data": meta["Data"],
                "Filename": filename,
                # Item
                "Item": orc_item["Item"],
                "Descricao": orc_item["Descricao"],
                "Quant": orc_item["Quant"],
                "Unid": orc_item["Unid"],
                "Tipo": orc_item["Tipo"],
                # Fornecedor
                "FornecedorIndex": forn["FornecedorIndex"],
                "FornecedorNome": forn["Nome"],
                "Contato": forn["Contato"],
                "Telefone": forn["Telefone"],
                "Email": forn["Email"],
                # Prices
                "ValorA": val_a,
                "ValorB": val_b,
                "Preco": preco,
            })

    df = pd.DataFrame(flat_records) if flat_records else pd.DataFrame()

    return {
        "meta": meta,
        "orcamento": orcamento,
        "fornecedores": [
            {k: v for k, v in f.items() if k != "StartCol"}
            for f in fornecedores
        ],
        "price_map": price_map,
        "flat": df,
    }


def parse_multiple_orcamento_files(
    files: List[Tuple[bytes, str]],
) -> pd.DataFrame:
    """
    Parse multiple Mapa de Concorrência files and merge into a single DataFrame.
    Each file adds its own items + fornecedores + prices.

    Args:
        files: list of (file_bytes, filename) tuples

    Returns:
        Combined DataFrame with all items × fornecedores across all files
    """
    all_dfs = []
    for file_bytes, filename in files:
        try:
            result = parse_orcamento_file(file_bytes, filename)
            if not result["flat"].empty:
                all_dfs.append(result["flat"])
        except Exception as e:
            print(f"Error parsing {filename}: {e}")
            continue

    if not all_dfs:
        return pd.DataFrame()

    return pd.concat(all_dfs, ignore_index=True)


# ─── Detection ────────────────────────────────────────────────────────────────

def detect_orcamento_file(file_bytes: bytes, filename: str) -> bool:
    """
    Detect if a file is a Mapa de Concorrência spreadsheet.
    Checks:
    1. Row 2 contains "MAPA DE CONCORRÊNCIA"
    2. Row 6 contains "FORNECEDOR"
    3. Row 13 contains "ITEM" and "DESCRIÇÃO"
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Check row 2 for title
        for col in range(1, 5):
            v = _safe_str(ws.cell(row=2, column=col).value).upper()
            if "MAPA" in v and "CONCORR" in v:
                wb.close()
                return True

        # Check row 13 for ITEM header
        v13 = _safe_str(ws.cell(row=13, column=2).value).upper()
        if "ITEM" in v13:
            wb.close()
            return True

        # Check row 6 for FORNECEDOR
        for col in range(1, 22):
            v = _safe_str(ws.cell(row=6, column=col).value).upper()
            if "FORNECEDOR" in v:
                wb.close()
                return True

        wb.close()
    except Exception:
        pass

    return False


# ─── CLI Test ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import json

    path = sys.argv[1] if len(sys.argv) > 1 else ""
    if not path:
        print("Usage: python orcamento_parser.py <file.xlsx>")
        sys.exit(1)

    with open(path, "rb") as f:
        content = f.read()

    result = parse_orcamento_file(content, path)

    print(f"Meta: {json.dumps(result['meta'], indent=2, ensure_ascii=False)}")
    print(f"\nOrçamento ({len(result['orcamento'])} items):")
    for item in result["orcamento"]:
        print(f"  {item['Item']:>3}. [{item['Tipo']:>7}] {item['Descricao'][:60]} | Q={item['Quant']} {item['Unid']}")
    print(f"\nFornecedores ({len(result['fornecedores'])}):")
    for f in result["fornecedores"]:
        print(f"  #{f['FornecedorIndex']} {f['Nome']} | {f['Telefone']} | {f['Email']}")
        print(f"       Headers: {f['PriceHeaderA']} / {f['PriceHeaderB']}")
    print(f"\nPrice Map ({len(result['price_map'])} entries):")
    for p in result["price_map"][:10]:
        print(f"  Item {p['Item']:>3} × {p['FornecedorNome'][:25]:25} → A={p['ValorA']}  B={p['ValorB']}  Preço={p['Preco']}")
    print(f"\nFlat DataFrame: {len(result['flat'])} rows × {len(result['flat'].columns)} cols")
    if not result["flat"].empty:
        print(result["flat"].head(5).to_string(index=False))

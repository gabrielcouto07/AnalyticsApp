"""
Materiais Template — Template definition for "Materiais/Mapa de Concorrência"
(Budget + Materials + Supplier Price Mapping) data.

This template handles construction material lists with:
- Multiple suppliers (Fornecedores) with pricing
- Material items with descriptions, quantities, units
- Price mapping for bidding/comparison
- Support for multiple tables/sheets within single Excel file

Detection: 
  - Row 2 contains "MAPA DE CONCORRÊNCIA" OR "MATERIAIS"
  - Row 13+ contains headers: ITEM, DESCRIÇÃO, QUANT., UNID., PREÇOS
  - Multiple tables supported (each supplier section)
  
Structure:
  - Can extract from sheets named: MP, MATERIALS, MATERIAIS
  - Detects supplier blocks by "FORNECEDOR N" pattern
  - Supports multiple sheets with different suppliers
"""

from typing import Dict, List, Any, Optional
from datetime import datetime

MATERIAIS_TEMPLATE = {
    "name": "Materiais - Mapa de Concorrência",
    "description": "Análise profissional de materiais e orçamento. Comparativo de preços entre fornecedores com cotações e contatos.",
    "icon": "🏗️",
    "color": "#4f8ef7",
    "tags": ["orçamento", "concorrência", "fornecedores", "preços"],
    "key_metrics": [
        {
            "name": "Total de Itens",
            "field": "Item",
            "description": "Quantidade única de materiais/itens",
            "type": "unique_count",
            "icon": "📦",
        },
        {
            "name": "Fornecedores Ativos",
            "field": "FornecedorNome",
            "description": "Número de fornecedores/orçamentistas",
            "type": "unique_count",
            "icon": "🤝",
        },
        {
            "name": "Valor Total Menor",
            "field": "ValorNegociado",
            "description": "Melhor orçamento total consolidado",
            "type": "min_sum",
            "icon": "💰",
        },
        {
            "name": "Projetos/Obras",
            "field": "Obra",
            "description": "Número de obras/projetos únicos",
            "type": "unique_count",
            "icon": "🏢",
        },
        {
            "name": "Valor Total em Análise",
            "field": "ValorTotal",
            "description": "Soma de todos os itens orçados",
            "type": "sum",
            "icon": "📊",
        },
    ],
    "required_columns": [
        "Obra",
        "Assunto",
        "Item",
        "Descricao",
        "Quant",
        "Unid",
        "FornecedorNome",
        "Contato",
        "Telefone",
        "Email",
        "ValorUnitario",
        "ValorNegociado",
        "ValorTotal",
    ],
    "visualizations": [
        {
            "type": "bar",
            "title": "Orçamento Total por Fornecedor",
            "field": "FornecedorNome",
            "value_field": "ValorTotal",
            "description": "Comparação de valor consolidado por fornecedor",
            "icon": "💸",
        },
        {
            "type": "table",
            "title": "Análise Comparativa de Preços",
            "pivot": True,
            "rows": "Descricao",
            "columns": "FornecedorNome",
            "values": "ValorUnitario",
            "description": "Matriz com preços unitários para decisão",
            "icon": "📊",
        },
        {
            "type": "bar",
            "title": "Quantidade de Itens por Assunto",
            "field": "Assunto",
            "value_field": "count",
            "description": "Distribuição de itens nos assuntos",
            "icon": "📋",
        },
        {
            "type": "pie",
            "title": "Distribuição de Valor por Assunto",
            "field": "Assunto",
            "value_field": "ValorTotal",
            "description": "Proporção do orçamento em cada assunto",
            "icon": "🥧",
        },
        {
            "type": "table",
            "title": "Rede de Contatos de Fornecedores",
            "fields": ["FornecedorNome", "Contato", "Telefone", "Email", "Endereco"],
            "description": "Lista completa de contatos para comunicação",
            "icon": "📞",
        },
        {
            "type": "heatmap",
            "title": "Competitividade de Preços",
            "x_field": "FornecedorNome",
            "y_field": "Assunto",
            "value_field": "ValorUnitario",
            "description": "Mapa de calor para identificar desvios de preço",
            "icon": "🔥",
        },
    ],
    "filters": [
        {"field": "Obra", "type": "single_select"},
        {"field": "Assunto", "type": "multi_select"},
        {"field": "FornecedorNome", "type": "multi_select"},
        {"field": "Unid", "type": "multi_select"},
        {"field": "ValorTotal", "type": "number_range"},
    ],
    "sample_columns": {
        "Obra": "text",
        "NumeroObra": "text",
        "Assunto": "text",
        "Item": "numeric",
        "Descricao": "text",
        "Quant": "numeric",
        "Unid": "text",
        "CodItem": "text",
        "FornecedorNumero": "numeric",
        "FornecedorNome": "text",
        "Contato": "text",
        "Telefone": "text",
        "Email": "text",
        "Endereco": "text",
        "ValorUnitario": "numeric",
        "ValorNegociado": "numeric",
        "ValorTotal": "numeric",
        "SaldoTotal": "numeric",
        "Data": "date",
    },
    "custom_parser": "materiais",
    "supports_multiple_tables": True,
    "table_detection": {
        "pattern": "MAPA DE CONCORRÊNCIA|MATERIAIS|FORNECEDOR",
        "sheet_names": ["MP", "MATERIALS", "MATERIAIS", "MAPA"],
        "headers_row": 13,
        "data_start_row": 17,
    },
    "parser_config": {
        "detect_suppliers": True,
        "merge_supplier_sheets": True,
        "flatten_to_rows": True,
        "expand_prices": True,
    },
}


def detect_materiais_file(file_bytes: bytes, filename: str) -> bool:
    """
    Detect if uploaded file is a Materiais/Mapa de Concorrência file.
    
    Checks for:
    - Filename patterns: 15.2.x, MP, Mapa, Materiais
    - Sheet names: MP, MATERIALS, MATERIAIS
    - Keywords in first rows: "MAPA DE CONCORRÊNCIA"
    """
    
    # Quick filename check first
    filename_lower = filename.lower()
    filename_indicators = ["15.2", "mapa", "materiais", "mp-", "ril-"]
    
    filename_match = any(ind in filename_lower for ind in filename_indicators)
    
    if not filename_match:
        return False
    
    # Verify with file content
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        
        # Check sheet name
        sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
        sheet_match = sheet_name.upper() in ["MP", "MATERIALS", "MATERIAIS", "MAPA"]
        
        ws = wb[sheet_name]
        
        # Check row 2 for "MAPA DE CONCORRÊNCIA"
        row_2_text = []
        for col in range(1, 10):
            cell_val = ws.cell(row=2, column=col).value
            if cell_val:
                row_2_text.append(str(cell_val).lower())
        
        row_2_combined = " ".join(row_2_text)
        keyword_match = "mapa" in row_2_combined and "concorrência" in row_2_combined
        
        wb.close()
        
        # Return True if sheet matches or keyword found
        return sheet_match or keyword_match
        
    except Exception:
        # If parsing fails but filename matched, still return True
        return True


def get_materiais_template() -> Dict[str, Any]:
    """Return the Materiais template definition."""
    return MATERIAIS_TEMPLATE

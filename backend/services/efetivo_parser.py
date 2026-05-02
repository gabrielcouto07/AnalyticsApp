from __future__ import annotations

import calendar
import re
import unicodedata
from io import BytesIO
from typing import Any

import openpyxl
import pandas as pd

from .data_quality import build_quality_report


MONTH_MAP = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARCOO": 3,
    "MARCO ": 3,
    "MARO": 3,
    "MARCO/": 3,
    "MARCO-": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

IGNORE_SHEETS = {"CONSULTA (ATIV - SERV)"}


def _normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = text.encode("ascii", "ignore").decode("ascii").upper()
    return re.sub(r"\s+", " ", text).strip()


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _extract_year_from_filename(filename: str) -> int:
    match = re.search(r"(20\d{2})", filename)
    return int(match.group(1)) if match else pd.Timestamp.today().year


def _find_month_number(sheet_name: str) -> int:
    normalized = _normalize_text(sheet_name)
    for token, month in MONTH_MAP.items():
        if token in normalized:
            return month
    return 1


def _find_day_header_row(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, list[tuple[int, int]]]:
    best_row = 0
    best_columns: list[tuple[int, int]] = []
    for row_idx in range(1, min(15, worksheet.max_row) + 1):
        current: list[tuple[int, int]] = []
        for col_idx in range(1, worksheet.max_column + 1):
            raw_value = worksheet.cell(row=row_idx, column=col_idx).value
            text = _safe_text(raw_value)
            if text.isdigit():
                day = int(text)
                if 1 <= day <= 31:
                    current.append((col_idx, day))
            elif isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool) and not pd.isna(raw_value):
                day = int(raw_value)
                if 1 <= day <= 31:
                    current.append((col_idx, day))
        if len(current) > len(best_columns):
            best_row = row_idx
            best_columns = current
    if len(best_columns) < 10:
        raise ValueError(f"Nao foi possivel detectar o cabecalho de dias na planilha {worksheet.title}.")
    return best_row, best_columns


def _classify_raw_value(value: Any) -> tuple[float | None, str, str | None]:
    if value is None:
        return None, "vazio", None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None, "vazio", None
        upper = text.upper()
        if upper in {"-", "–", "—", "â€“", "â€”"}:
            return None, "traco", None
        if upper == "NA":
            return None, "na", None
        if text.startswith("#"):
            return None, "erro", text
        cleaned = (
            text.replace("R$", "")
            .replace("\xa0", " ")
            .replace(".", "")
            .replace(",", ".")
            .strip()
        )
        cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
        if cleaned:
            try:
                numeric = float(cleaned)
            except ValueError:
                return None, "vazio", text
            if numeric == 0:
                return 0.0, "zero", None
            return numeric, "numero", None
        return None, "vazio", text
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if pd.isna(value):
            return None, "vazio", None
        numeric = float(value)
        if numeric == 0:
            return 0.0, "zero", None
        return numeric, "numero", None
    return None, "vazio", str(value)


def _is_blank_day_value(value: Any) -> bool:
    return _classify_raw_value(value)[1] == "vazio"


def _extract_obra_name(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    for row_idx in range(1, min(5, worksheet.max_row) + 1):
        values = [
            _safe_text(worksheet.cell(row_idx, col_idx).value)
            for col_idx in range(1, min(6, worksheet.max_column) + 1)
        ]
        non_empty = [value for value in values if value]
        if non_empty:
            return non_empty[0]
    return worksheet.title


def _parse_consulta_lookup(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    atividades: list[str] = []
    servicos: list[str] = []
    insumos: list[str] = []
    classificacoes: list[dict[str, str]] = []

    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        tipo = _safe_text(values[0] if len(values) > 0 else "")
        descricao = _safe_text(values[1] if len(values) > 1 else "")
        atividade = _safe_text(values[3] if len(values) > 3 else "")
        servico = _safe_text(values[5] if len(values) > 5 else "")
        insumo = _safe_text(values[7] if len(values) > 7 else "")
        if tipo and descricao:
            classificacoes.append({"tipo": tipo, "descricao": descricao})
        if atividade:
            atividades.append(atividade)
        if servico:
            servicos.append(servico)
        if insumo:
            insumos.append(insumo)

    return {
        "atividades": sorted(set(atividades)),
        "servicos": sorted(set(servicos)),
        "insumos": sorted(set(insumos)),
        "classificacoes": classificacoes,
    }


def parse_efetivo_sheet(df_raw: pd.DataFrame, sheet_name: str, year: int | None = None) -> pd.DataFrame:
    if df_raw.empty:
        return pd.DataFrame()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in df_raw.itertuples(index=False):
        worksheet.append(list(row))

    resolved_year = year or pd.Timestamp.today().year
    obra = _extract_obra_name(worksheet)
    header_row, day_columns = _find_day_header_row(worksheet)
    month_num = _find_month_number(sheet_name)
    max_days = calendar.monthrange(resolved_year, month_num)[1]
    day_columns = [(col_idx, day) for col_idx, day in day_columns if day <= max_days]
    if not day_columns:
        return pd.DataFrame()

    current_supplier = ""
    records: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row_label = _safe_text(worksheet.cell(row=row_idx, column=1).value)
        if not row_label:
            continue

        raw_day_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx, _ in day_columns]
        blank_ratio = sum(1 for value in raw_day_values if _is_blank_day_value(value)) / max(len(raw_day_values), 1)
        has_any_non_blank = any(not _is_blank_day_value(value) for value in raw_day_values)
        normalized_label = _normalize_text(row_label)

        if normalized_label.startswith("TOTAL"):
            continue

        if blank_ratio >= 0.8 and not has_any_non_blank:
            current_supplier = row_label
            continue

        fornecedor = current_supplier or row_label
        funcao = row_label
        for col_idx, day in day_columns:
            raw_value = worksheet.cell(row=row_idx, column=col_idx).value
            quantidade, tipo_valor, observacao = _classify_raw_value(raw_value)
            data = pd.Timestamp(year=resolved_year, month=month_num, day=day)
            records.append(
                {
                    "arquivo_origem": "",
                    "obra": obra,
                    "mes": sheet_name,
                    "mes_num": month_num,
                    "dia": day,
                    "data": data,
                    "fornecedor": fornecedor,
                    "funcao": funcao,
                    "quantidade_efetivo": quantidade,
                    "valor_original_celula": raw_value,
                    "tipo_valor": tipo_valor,
                    "observacao": observacao,
                    "source_sheet": sheet_name,
                }
            )

    return pd.DataFrame(records)


def _worksheet_to_raw_dataframe(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> pd.DataFrame:
    rows = list(worksheet.iter_rows(values_only=True))
    return pd.DataFrame(rows)


def _build_legacy_columns(dataframe: pd.DataFrame, filename: str, year: int) -> pd.DataFrame:
    if dataframe.empty:
        return dataframe

    working = dataframe.copy()
    working["arquivo_origem"] = filename
    working["Obra"] = working["obra"]
    working["Ano"] = year
    working["Mes"] = working["mes_num"]
    working["MesNome"] = working["mes"]
    working["Fornecedor"] = working["fornecedor"]
    working["Funcao"] = working["funcao"]
    working["Quantidade"] = pd.to_numeric(working["quantidade_efetivo"], errors="coerce").fillna(0)
    working["Dia"] = working["dia"]
    working["Data"] = pd.to_datetime(working["data"], errors="coerce")
    working["Periodo"] = working["Data"].dt.strftime("%Y-%m")
    working["DiaSemana"] = working["Data"].dt.day_name()
    working["FornecedorFuncao"] = working["Fornecedor"].fillna("") + " | " + working["Funcao"].fillna("")
    working["Trabalhou"] = (working["Quantidade"] > 0).astype(int)
    group_totals = (
        working[working["tipo_valor"] == "numero"]
        .groupby(["Fornecedor", "Funcao", "Mes"], dropna=False)["Quantidade"]
        .sum()
        .rename("DiariasTotal")
        .reset_index()
    )
    working = working.merge(group_totals, on=["Fornecedor", "Funcao", "Mes"], how="left")
    working["DiariasTotal"] = working["DiariasTotal"].fillna(0)
    return working


def _build_aggregated_payload(dataframe: pd.DataFrame) -> dict[str, Any]:
    numeric_records = (
        dataframe[dataframe["tipo_valor"] == "numero"].copy()
        if not dataframe.empty and "tipo_valor" in dataframe.columns
        else pd.DataFrame()
    )
    if numeric_records.empty:
        return {
            "por_mes": [],
            "por_fornecedor": [],
            "por_funcao": [],
            "total_diarias": 0.0,
            "media_diaria": 0.0,
            "pico_diario": 0.0,
            "dias_ativos": 0,
            "fornecedores_ativos": [],
            "funcoes_ativas": [],
        }

    numeric_records["quantidade_efetivo"] = pd.to_numeric(
        numeric_records["quantidade_efetivo"],
        errors="coerce",
    ).fillna(0)
    daily_totals = (
        numeric_records.groupby("data", dropna=False)["quantidade_efetivo"]
        .sum()
        .reset_index(name="total")
    )
    return {
        "por_mes": (
            numeric_records.groupby(["mes_num", "mes"], dropna=False)["quantidade_efetivo"]
            .sum()
            .reset_index()
            .sort_values(["mes_num", "mes"])
            .rename(columns={"quantidade_efetivo": "total"})
            .to_dict(orient="records")
        ),
        "por_fornecedor": (
            numeric_records.groupby("fornecedor", dropna=False)["quantidade_efetivo"]
            .sum()
            .reset_index()
            .sort_values("quantidade_efetivo", ascending=False)
            .rename(columns={"quantidade_efetivo": "total"})
            .to_dict(orient="records")
        ),
        "por_funcao": (
            numeric_records.groupby("funcao", dropna=False)["quantidade_efetivo"]
            .sum()
            .reset_index()
            .sort_values("quantidade_efetivo", ascending=False)
            .rename(columns={"quantidade_efetivo": "total"})
            .to_dict(orient="records")
        ),
        "total_diarias": round(float(numeric_records["quantidade_efetivo"].sum()), 2),
        "media_diaria": round(float(daily_totals["total"].mean()), 2) if not daily_totals.empty else 0.0,
        "pico_diario": round(float(daily_totals["total"].max()), 2) if not daily_totals.empty else 0.0,
        "dias_ativos": int(daily_totals[daily_totals["total"] > 0]["data"].nunique()) if not daily_totals.empty else 0,
        "fornecedores_ativos": sorted(numeric_records["fornecedor"].dropna().astype(str).unique().tolist()),
        "funcoes_ativas": sorted(numeric_records["funcao"].dropna().astype(str).unique().tolist()),
    }


def parse_efetivo_workbook(file_bytes: bytes, filename: str) -> dict[str, Any]:
    year = _extract_year_from_filename(filename)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    parsed_frames: list[pd.DataFrame] = []
    raw_sheets: dict[str, pd.DataFrame] = {}
    consulta_lookup: dict[str, Any] = {}
    normalization_notes: list[str] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        raw_df = _worksheet_to_raw_dataframe(worksheet)
        raw_sheets[sheet_name] = raw_df

        if _normalize_text(sheet_name) in {_normalize_text(name) for name in IGNORE_SHEETS}:
            consulta_lookup = _parse_consulta_lookup(worksheet)
            continue

        parsed = parse_efetivo_sheet(raw_df, sheet_name, year=year)
        if parsed.empty:
            continue
        parsed["arquivo_origem"] = filename
        parsed_frames.append(parsed)
        normalization_notes.append(f"Summary columns for '{sheet_name}' were excluded from daily records.")

    merged = pd.concat(parsed_frames, ignore_index=True) if parsed_frames else pd.DataFrame()
    merged = _build_legacy_columns(merged, filename, year)
    aggregated = _build_aggregated_payload(merged)
    quality_report = build_quality_report(
        raw_sheets,
        normalization_notes=normalization_notes,
    ).to_dict()

    metadata = {
        "obra": merged["obra"].dropna().astype(str).iloc[0] if not merged.empty and "obra" in merged.columns else "",
        "year": year,
        "filename": filename,
    }
    return {
        "records": merged,
        "consulta_lookup": consulta_lookup,
        "metadata": metadata,
        "aggregated": aggregated,
        "quality_report": quality_report,
    }


def parse_efetivo_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    return parse_efetivo_workbook(file_bytes, filename)["records"]


def get_efetivo_summary(df: pd.DataFrame) -> dict[str, Any]:
    if df.empty:
        return {}
    working = df.copy()
    if "tipo_valor" in working.columns:
        working = working[working["tipo_valor"] == "numero"].copy()
    working["quantidade_efetivo"] = pd.to_numeric(working.get("quantidade_efetivo"), errors="coerce").fillna(0)
    return {
        "total_diarias": round(float(working["quantidade_efetivo"].sum()), 2),
        "fornecedores": sorted(working.get("fornecedor", pd.Series(dtype=object)).dropna().astype(str).unique().tolist()),
        "funcoes": sorted(working.get("funcao", pd.Series(dtype=object)).dropna().astype(str).unique().tolist()),
        "meses": sorted(working.get("mes", pd.Series(dtype=object)).dropna().astype(str).unique().tolist()),
    }

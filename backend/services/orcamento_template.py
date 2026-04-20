"""
Orcamento Template — Template definition for "Mapa de Concorrência"
(Budget + Supplier + Price Map) data.

These files track:
- Orçamento: items with description, quantity, unit
- Fornecedores: supplier name, contact, phone, email
- Price Map: each fornecedor's pricing per item (two price columns)

Detection: row 2 has "MAPA DE CONCORRÊNCIA", row 6 has "FORNECEDOR N",
           row 13 has "ITEM" + "DESCRIÇÃO".
"""

from typing import Dict, List, Any

ORCAMENTO_TEMPLATE = {
    "name": "Orçamento - Mapa de Concorrência",
    "description": "Análise competitiva de orçamento. Comparação de preços por fornecedor com classificação de serviços e insumos.",
    "icon": "💰",
    "color": "#10b981",
    "tags": ["orçamento", "concorrência", "preços", "fornecedores"],
    "key_metrics": [
        {
            "name": "Total de Itens",
            "field": "Item",
            "description": "Quantidade total de linhas orçamentárias",
            "type": "unique_count",
            "icon": "📋",
        },
        {
            "name": "Fornecedores Cotados",
            "field": "FornecedorNome",
            "description": "Número de fornecedores em análise",
            "type": "unique_count",
            "icon": "🤝",
        },
        {
            "name": "Melhor Preço Total",
            "field": "Preco",
            "description": "Orçamento consolidado mais vantajoso",
            "type": "min_sum",
            "icon": "✅",
        },
        {
            "name": "Classificação de Itens",
            "field": "Tipo",
            "description": "Distribuição entre serviços e insumos",
            "type": "count_group",
            "icon": "🏷️",
        },
        {
            "name": "Valor Total em Disputa",
            "field": "Preco",
            "description": "Soma de todos os itens cotados",
            "type": "sum",
            "icon": "💵",
        },
    ],
    "required_columns": [
        "Obra",
        "Assunto",
        "Item",
        "Descricao",
        "Quant",
        "Unid",
        "Tipo",
        "FornecedorNome",
        "Preco",
    ],
    "visualizations": [
        {
            "type": "bar",
            "title": "Preço Total Consolidado por Fornecedor",
            "field": "FornecedorNome",
            "value_field": "Preco",
            "description": "Comparação de orçamento total para melhor decisão",
            "icon": "📊",
        },
        {
            "type": "table",
            "title": "Matriz Comparativa de Preços",
            "pivot": True,
            "rows": "Descricao",
            "columns": "FornecedorNome",
            "values": "Preco",
            "description": "Análise item-a-item para identificar diferenciais",
            "icon": "📈",
        },
        {
            "type": "pie",
            "title": "Composição: Serviços vs Insumos",
            "field": "Tipo",
            "value_field": "count",
            "description": "Proporção de itens em cada categoria",
            "icon": "🥧",
        },
        {
            "type": "bar",
            "title": "Distribuição de Valor por Tipo",
            "field": "Tipo",
            "value_field": "Preco",
            "description": "Quanto é investido em cada tipo",
            "icon": "💰",
        },
        {
            "type": "bar",
            "title": "Análise de Assuntos/Projetos",
            "field": "Assunto",
            "value_field": "Preco",
            "description": "Segregação de custos por projeto",
            "icon": "🏢",
        },
        {
            "type": "heatmap",
            "title": "Variação de Preços: Fornecedor vs Tipo",
            "x_field": "FornecedorNome",
            "y_field": "Tipo",
            "value_field": "Preco",
            "description": "Identifique onde cada fornecedor é mais competitivo",
            "icon": "🔥",
        },
    ],
    "filters": [
        {"field": "FornecedorNome", "type": "multi_select"},
        {"field": "Tipo", "type": "multi_select"},
        {"field": "Assunto", "type": "multi_select"},
        {"field": "Preco", "type": "number_range"},
    ],
    "sample_columns": {
        "Obra": "text",
        "Assunto": "text",
        "Item": "numeric",
        "Descricao": "text",
        "Quant": "numeric",
        "Unid": "text",
        "Tipo": "text",
        "FornecedorNome": "text",
        "Contato": "text",
        "Telefone": "text",
        "Email": "text",
        "ValorA": "numeric",
        "ValorB": "numeric",
        "Preco": "numeric",
    },
    "custom_parser": "orcamento",
}


ORCAMENTO_DETECTION_KEYWORDS = {
    "strong": [
        "mapa de concorrência", "descrição do serviço", "fornecedor",
        "valor unit", "valor total", "valor negociado", "valor inicial",
        "saldo total",
    ],
    "medium": [
        "item", "quant", "unid", "preços", "concorrência",
    ],
    "weak": [
        "obra", "assunto", "orçamento",
    ],
}


def detect_orcamento_file(file_bytes: bytes, filename: str) -> bool:
    """
    Detect if a file is a Mapa de Concorrência spreadsheet.
    Checks row 2 for "MAPA DE CONCORRÊNCIA" and row 6/13 for structural markers.
    """
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Row 2: title
        for col in range(1, 5):
            v = str(ws.cell(row=2, column=col).value or "").upper()
            if "MAPA" in v and "CONCORR" in v:
                wb.close()
                return True

        # Row 13: ITEM header
        v13 = str(ws.cell(row=13, column=2).value or "").upper()
        if "ITEM" in v13:
            wb.close()
            return True

        # Row 6: FORNECEDOR
        for col in range(1, 22):
            v = str(ws.cell(row=6, column=col).value or "").upper()
            if "FORNECEDOR" in v:
                wb.close()
                return True

        wb.close()
    except Exception:
        pass
    return False


def get_orcamento_template() -> Dict[str, Any]:
    return ORCAMENTO_TEMPLATE


def get_orcamento_suggestions_score(columns: List[str]) -> float:
    col_lower = [c.lower().strip() for c in columns]
    score = 0.0
    for col in col_lower:
        for kw in ORCAMENTO_DETECTION_KEYWORDS["strong"]:
            if kw in col or col in kw:
                score += 3.0
        for kw in ORCAMENTO_DETECTION_KEYWORDS["medium"]:
            if kw in col or col in kw:
                score += 1.5
        for kw in ORCAMENTO_DETECTION_KEYWORDS["weak"]:
            if kw in col or col in kw:
                score += 1.0
    return score

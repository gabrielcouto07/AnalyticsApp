"""Excel → SQL + Excel-formula → JavaScript conversion."""
import io
import re
from typing import Any
import pandas as pd
from openpyxl import load_workbook


# ---------- SQL generation ----------

def _sql_type(dtype: Any) -> str:
    s = str(dtype)
    if "int" in s:
        return "INTEGER"
    if "float" in s:
        return "DECIMAL(18, 4)"
    if "bool" in s:
        return "BOOLEAN"
    if "datetime" in s:
        return "TIMESTAMP"
    return "TEXT"


def _safe_ident(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", str(name)).strip("_")
    if not cleaned:
        cleaned = "col"
    if cleaned[0].isdigit():
        cleaned = "_" + cleaned
    return cleaned.lower()


def _sql_literal(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, pd.Timestamp):
        return f"'{v.isoformat()}'"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def build_sql(df: pd.DataFrame, table_name: str = "imported_data", row_limit: int = 50) -> dict:
    table = _safe_ident(table_name)
    cols = [(c, _safe_ident(c), _sql_type(df[c].dtype)) for c in df.columns]

    # CREATE TABLE
    create_lines = [f"CREATE TABLE {table} ("]
    for i, (orig, safe, sql_type) in enumerate(cols):
        comma = "," if i < len(cols) - 1 else ""
        create_lines.append(f"    {safe} {sql_type}{comma}  -- {orig}")
    create_lines.append(");")
    create_table = "\n".join(create_lines)

    # One ALTER TABLE per column (matches "for each column, a SQL command to add it")
    alter_columns = [
        {
            "column": orig,
            "sql": f"ALTER TABLE {table} ADD COLUMN {safe} {sql_type};",
        }
        for orig, safe, sql_type in cols
    ]

    # Sample INSERTs (capped to keep payload small)
    safe_cols = ", ".join(safe for _, safe, _ in cols)
    inserts = []
    for _, row in df.head(row_limit).iterrows():
        values = ", ".join(_sql_literal(row[c]) for c, _, _ in cols)
        inserts.append(f"INSERT INTO {table} ({safe_cols}) VALUES ({values});")

    return {
        "table_name": table,
        "create_table": create_table,
        "alter_columns": alter_columns,
        "inserts": inserts,
        "rows_inserted": len(inserts),
        "rows_total": len(df),
    }


# ---------- Excel formula → JavaScript ----------

_CELL_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")
_RANGE_RE = re.compile(r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)")
_FUNC_RE = re.compile(r"\b([A-Z][A-Z0-9_]*)\s*\(")
_STRING_RE = re.compile(r'"([^"]*)"')


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _index_to_col_letters(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _expand_range(match: re.Match) -> str:
    c1, r1, c2, r2 = match.group(1), int(match.group(2)), match.group(3), int(match.group(4))
    i1, i2 = _col_letters_to_index(c1), _col_letters_to_index(c2)
    cells = []
    for ci in range(min(i1, i2), max(i1, i2) + 1):
        for ri in range(min(r1, r2), max(r1, r2) + 1):
            cells.append(f"{_index_to_col_letters(ci).lower()}{ri}")
    return "[" + ", ".join(cells) + "]"


_FUNC_TEMPLATES = {
    "SUM":         "({arg}).reduce((a,b)=>a+(+b||0),0)",
    "AVERAGE":     "(({arg}).reduce((a,b)=>a+(+b||0),0)/({arg}).length)",
    "AVG":         "(({arg}).reduce((a,b)=>a+(+b||0),0)/({arg}).length)",
    "COUNT":       "({arg}).filter(v=>v!==null&&v!==undefined&&v!=='').length",
    "MIN":         "Math.min(...[].concat({arg}))",
    "MAX":         "Math.max(...[].concat({arg}))",
    "ABS":         "Math.abs({arg})",
    "SQRT":        "Math.sqrt({arg})",
    "ROUND":       "Math.round({arg})",
    "FLOOR":       "Math.floor({arg})",
    "CEILING":     "Math.ceil({arg})",
    "POWER":       "Math.pow({arg})",
    "MOD":         "(({arg})[0] % ({arg})[1])",
    "CONCAT":      "[{arg}].join('')",
    "CONCATENATE": "[{arg}].join('')",
    "LEN":         "String({arg}).length",
    "UPPER":       "String({arg}).toUpperCase()",
    "LOWER":       "String({arg}).toLowerCase()",
    "TRIM":        "String({arg}).trim()",
    "AND":         "[{arg}].every(Boolean)",
    "OR":          "[{arg}].some(Boolean)",
    "NOT":         "!({arg})",
    "TRUE":        "true",
    "FALSE":       "false",
}


def _split_args(s: str) -> list[str]:
    """Split a function arg list on top-level commas (respects parens and quoted strings)."""
    args, cur, depth, in_str = [], "", 0, False
    for ch in s:
        if ch == '"':
            in_str = not in_str
            cur += ch
        elif not in_str and ch == "(":
            depth += 1
            cur += ch
        elif not in_str and ch == ")":
            depth -= 1
            cur += ch
        elif not in_str and ch == "," and depth == 0:
            args.append(cur.strip())
            cur = ""
        else:
            cur += ch
    if cur.strip():
        args.append(cur.strip())
    return args


def _convert_functions(expr: str, unknown: set) -> str:
    """Replace innermost function calls until none remain."""
    while True:
        target = None
        for m in _FUNC_RE.finditer(expr):
            paren_start = m.end() - 1
            depth, i = 1, paren_start + 1
            while i < len(expr) and depth > 0:
                if expr[i] == "(":
                    depth += 1
                elif expr[i] == ")":
                    depth -= 1
                i += 1
            if depth != 0:
                continue
            inner = expr[paren_start + 1 : i - 1]
            if not _FUNC_RE.search(inner):  # only innermost
                target = (m, paren_start, i, inner)
                break

        if not target:
            break

        m, _start, end, inner = target
        fn = m.group(1).upper()
        args = _split_args(inner)
        joined = ", ".join(args)

        if fn == "IF":
            if len(args) == 3:
                replacement = f"(({args[0]}) ? ({args[1]}) : ({args[2]}))"
            elif len(args) == 2:
                replacement = f"(({args[0]}) ? ({args[1]}) : null)"
            else:
                replacement = f"({joined})"
        elif fn == "IFERROR":
            if len(args) == 2:
                replacement = f"(()=>{{ try {{ return ({args[0]}); }} catch(_) {{ return ({args[1]}); }} }})()"
            else:
                replacement = f"({joined})"
        elif fn in _FUNC_TEMPLATES:
            replacement = _FUNC_TEMPLATES[fn].format(arg=joined)
        else:
            unknown.add(fn)
            replacement = f"{fn.lower()}({joined})"

        expr = expr[: m.start()] + replacement + expr[end:]
    return expr


def excel_formula_to_js(formula: str) -> dict:
    """Convert an Excel formula string into a JavaScript expression.

    Returns: {"js": str, "unknown_functions": list[str]}
    Returns the original string if the input isn't a formula.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return {"js": formula, "unknown_functions": []}

    expr = formula[1:].strip()

    # Mask string literals so cell-ref / operator regexes don't touch their contents
    strings: list[str] = []

    def _mask(m):
        strings.append(m.group(0))
        return f"\x00{len(strings) - 1}\x00"

    expr = _STRING_RE.sub(_mask, expr)

    # Order matters: ranges → functions → cell refs → operators
    expr = _RANGE_RE.sub(_expand_range, expr)
    unknown: set = set()
    expr = _convert_functions(expr, unknown)
    expr = _CELL_RE.sub(lambda m: f"{m.group(1).lower()}{m.group(2)}", expr)

    expr = expr.replace("<>", "!==")
    expr = expr.replace("^", "**")
    # Excel `=` is equality. Skip `=>` (JS arrow funcs we just generated) and `==`.
    expr = re.sub(r"(?<![=<>!])=(?![=>])", " === ", expr)
    # Excel `&` is string concat. Skip `&&` so we don't mangle JS we generated.
    expr = re.sub(r"&(?!&)", " + ", expr)

    # Restore strings
    expr = re.sub(r"\x00(\d+)\x00", lambda m: strings[int(m.group(1))], expr)

    return {"js": expr.strip(), "unknown_functions": sorted(unknown)}


def extract_formulas(file_bytes: bytes) -> list[dict]:
    """Walk every sheet and return all formula cells with their JS translation."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    out = []
    for ws in wb.worksheets:
        headers: dict[int, str] = {}
        first = next(ws.iter_rows(min_row=1, max_row=1), [])
        for cell in first:
            if cell.value is not None:
                headers[cell.column] = str(cell.value)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    converted = excel_formula_to_js(v)
                    out.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "column": headers.get(cell.column, ""),
                        "excel": v,
                        "javascript": converted["js"],
                        "unknown_functions": converted["unknown_functions"],
                    })
    return out


def analyze_xlsx(file_bytes: bytes, filename: str) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes))

    columns = []
    for c in df.columns:
        non_null = df[c].dropna()
        columns.append({
            "original": str(c),
            "safe": _safe_ident(c),
            "sql_type": _sql_type(df[c].dtype),
            "pandas_dtype": str(df[c].dtype),
            "sample": str(non_null.iloc[0]) if not non_null.empty else "",
            "non_null_count": int(non_null.shape[0]),
        })

    table_base = re.sub(r"\.[^.]+$", "", filename or "imported_data")
    sql = build_sql(df, table_name=table_base)
    formulas = extract_formulas(file_bytes)

    return {
        "filename": filename,
        "rows": int(len(df)),
        "column_count": int(len(df.columns)),
        "columns": columns,
        "sql": sql,
        "formulas": formulas,
        "formula_count": len(formulas),
    }

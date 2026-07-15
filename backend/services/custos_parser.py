"""
Custos Parser — Extracts data from "Planilha Controle de Custos Consolidados" (.xlsm/.xlsx).

Performance note: For .xlsm files, openpyxl read_only=True is slower on row iteration
(10s for 300 rows) while read_only=False loads in ~8s but iterates instantly.
We use read_only=False since total time is similar but more predictable.

Two tables:
1. "PLANILHA NFs - Entrada de Dados" (row 9 headers, row 10+ data)
   Stops when FORNECEDOR (col 4) is empty for 15 consecutive rows.
2. "PLANILHA CONSOLIDADO" (row 7 headers, row 9+ data)
   Stops when FORNECEDOR (col 3) is '---' or empty for 15 consecutive rows.
"""

import re
import logging
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from io import BytesIO
import openpyxl

logger = logging.getLogger(__name__)


def _safe_str(v) -> str:
    if v is None: return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan") else s


def _parse_num(v) -> Optional[float]:
    if v is None: return None
    if isinstance(v, (int, float)):
        return None if (np.isnan(v) or np.isinf(v)) else float(v)
    s = str(v).strip()
    if s in ("", "-", "—", "nan", "None", "#REF!", "#VALUE!"): return None
    s = re.sub(r"[R$\s]", "", s)
    if re.search(r"\d\.\d{3},\d", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try: return float(s)
    except: return None


def _parse_date(v):
    if v is None: return None
    if hasattr(v, "strftime"): return v
    s = str(v).strip()
    if s.upper() in ("", "PAGO", "---", "NONE", "NAN"): return None
    try: return pd.to_datetime(s, dayfirst=True)
    except: return None


# ─── Column Maps ──────────────────────────────────────────────────────────────

NF_COL_MAP = {
    2: "NumConsolidado", 3: "Cod", 4: "Fornecedor", 5: "NF",
    7: "MapaPrecos", 8: "Natureza", 9: "CondPagto", 10: "DataVencto",
    11: "Valor", 12: "ItemPlanilha", 13: "ValorItem",
    15: "SituacaoPlanilha", 16: "SituacaoMapaCompra", 17: "SaldoPlanilha",
    18: "Observacoes", 19: "DescricaoItem",
}
NF_NUMERIC = {"Valor", "ValorItem", "SaldoPlanilha"}
NF_DATE = {"DataVencto"}

CONS_COL_MAP = {
    2: "NumConsolidado", 3: "Fornecedor", 4: "NF", 5: "Mapa",
    6: "Natureza", 7: "CondPagto", 8: "DataVencto",
    9: "Valor", 10: "ApropriItem", 11: "ApropriValor",
}
CONS_NUMERIC = {"Valor", "ApropriValor"}
CONS_DATE = {"DataVencto"}


def _parse_sheet(ws, col_map: dict, numeric_cols: set, date_cols: set,
                 data_start: int, forn_col: int, max_empty: int = 15,
                 skip_values: set = None) -> pd.DataFrame:
    """Generic sheet parser with empty-row detection."""
    records = []
    empty_streak = 0
    skip_values = skip_values or set()

    for row in range(data_start, ws.max_row + 1):
        forn = _safe_str(ws.cell(row=row, column=forn_col).value)

        if not forn or forn in skip_values:
            empty_streak += 1
            if empty_streak >= max_empty:
                break
            continue
        else:
            empty_streak = 0

        record = {}
        for col, name in col_map.items():
            v = ws.cell(row=row, column=col).value
            if name in numeric_cols:
                record[name] = _parse_num(v)
            elif name in date_cols:
                record[name] = _parse_date(v)
            else:
                record[name] = _safe_str(v)
        records.append(record)

    df = pd.DataFrame(records)
    if not df.empty:
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def _extract_meta(ws) -> Dict[str, str]:
    obra = _safe_str(ws.cell(row=7, column=2).value)
    periodo = _safe_str(ws.cell(row=7, column=4).value)
    endereco = _safe_str(ws.cell(row=8, column=2).value)
    total_nfs = _parse_num(ws.cell(row=8, column=8).value)
    total_valor = _parse_num(ws.cell(row=8, column=11).value)
    return {
        "Obra": obra.replace("OBRA:", "").strip() if obra else "",
        "Periodo": periodo,
        "Endereco": endereco.replace("ENDEREÇO:", "").strip() if endereco else "",
        "TotalNFs": int(total_nfs) if total_nfs else 0,
        "TotalValor": round(total_valor, 2) if total_valor else 0,
    }


# ─── Dynamic Header Detection ─────────────────────────────────────────────────

# Maps header keywords → canonical field names (order matters — first match wins)
_NF_HEADER_MAP = [
    ("CONSOLIDADO", "NumConsolidado"),
    ("FORNECEDOR",  "Fornecedor"),
    ("COD",         "Cod"),
    ("MAPA",        "MapaPrecos"),
    ("NATUREZA",    "Natureza"),
    ("COND",        "CondPagto"),
    ("BOLETO",      "CondPagto"),    # header appears as "BOLETO/DEPÓSITO" in real files
    ("VENCTO",      "DataVencto"),
    ("VENCIMENTO",  "DataVencto"),
    ("VALOR",       "Valor"),
    ("ITEM",        "ItemPlanilha"),
    ("NF",          "NF"),           # keep last — "NF" is a substring of many words
]

_CONS_HEADER_MAP = [
    ("CONSOLIDADO", "NumConsolidado"),
    ("FORNECEDOR",  "Fornecedor"),
    ("MAPA",        "Mapa"),
    ("NATUREZA",    "Natureza"),
    ("COND",        "CondPagto"),
    ("BOLETO",      "CondPagto"),    # header appears as "BOLETO/DEPÓSITO" in real files
    ("VENCTO",      "DataVencto"),
    ("VENCIMENTO",  "DataVencto"),
    ("VALOR",       "Valor"),
    ("APROPI",      "ApropriValor"),
    ("NF",          "NF"),
]


def _detect_layout(ws, anchor_keyword: str, header_map: list,
                   max_scan_rows: int = 20, max_cols: int = 40) -> tuple:
    """
    Scan the sheet for a row containing `anchor_keyword`.
    Returns (data_start_row, col_map, forn_col) using the header_map to name columns.
    Returns (None, {}, None) if not found.
    """
    for row in range(1, max_scan_rows + 1):
        for col in range(1, max_cols + 1):
            cell = _safe_str(ws.cell(row=row, column=col).value).upper()
            if anchor_keyword in cell:
                # Found the anchor — now map every column in this header row
                col_map = {}
                seen = set()
                for c in range(1, max_cols + 1):
                    hdr = _safe_str(ws.cell(row=row, column=c).value).upper()
                    if not hdr:
                        continue
                    for keyword, field_name in header_map:
                        if keyword in hdr and field_name not in seen:
                            col_map[c] = field_name
                            seen.add(field_name)
                            break
                forn_col = next((c for c, f in col_map.items() if f == "Fornecedor"), None)
                logger.info(
                    f"[custos_parser] layout detected: anchor='{anchor_keyword}' at row={row}, "
                    f"forn_col={forn_col}, map={col_map}"
                )
                return row + 1, col_map, forn_col
    return None, {}, None


def parse_custos_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    Parse Planilha Controle de Custos. Returns {meta, nfs, consolidado}.
    Uses read_only=False for predictable performance on .xlsm files.
    Performs dynamic header-row detection so column offsets in the real file
    are handled automatically — hardcoded fallback (row 10, col 4) only kicks
    in if "FORNECEDOR" cannot be found in the first 20 rows.
    """
    logger.info(f"[custos_parser] Starting parse of {filename} ({len(file_bytes)/1024:.0f}KB)")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)

    # ── NFs sheet ──────────────────────────────────────────────────────
    nfs_sheet = (
        next((s for s in wb.sheetnames if "NF" in s.upper() and "ENTRADA" in s.upper()), None)
        or next((s for s in wb.sheetnames if "NF" in s.upper()), None)
    )
    logger.info(f"[custos_parser] sheets={wb.sheetnames}, nfs_sheet={nfs_sheet}")
    df_nfs = pd.DataFrame()
    meta = {}

    if nfs_sheet:
        ws_nfs = wb[nfs_sheet]
        meta = _extract_meta(ws_nfs)

        data_start, col_map, forn_col = _detect_layout(
            ws_nfs, "FORNECEDOR", _NF_HEADER_MAP
        )

        if data_start and col_map and forn_col:
            df_nfs = _parse_sheet(
                ws_nfs, col_map, NF_NUMERIC, NF_DATE,
                data_start=data_start, forn_col=forn_col, max_empty=15,
            )
        else:
            logger.warning("[custos_parser] NFS header not found — using hardcoded fallback (row 10, col 4)")
            df_nfs = _parse_sheet(
                ws_nfs, NF_COL_MAP, NF_NUMERIC, NF_DATE,
                data_start=10, forn_col=4, max_empty=15,
            )

    # ── Consolidado sheet ──────────────────────────────────────────────
    cons_sheet = (
        next((s for s in wb.sheetnames if "CONSOLIDADO" in s.upper() and "RESUMO" not in s.upper()), None)
        or next((s for s in wb.sheetnames if "CONSOLIDADO" in s.upper()), None)
    )
    df_cons = pd.DataFrame()

    if cons_sheet:
        ws_cons = wb[cons_sheet]
        data_start_c, col_map_c, forn_col_c = _detect_layout(
            ws_cons, "FORNECEDOR", _CONS_HEADER_MAP
        )
        if data_start_c and col_map_c and forn_col_c:
            df_cons = _parse_sheet(
                ws_cons, col_map_c, CONS_NUMERIC, CONS_DATE,
                data_start=data_start_c, forn_col=forn_col_c,
                max_empty=15, skip_values={"---"},
            )
        else:
            logger.warning("[custos_parser] CONS header not found — using hardcoded fallback (row 9, col 3)")
            df_cons = _parse_sheet(
                ws_cons, CONS_COL_MAP, CONS_NUMERIC, CONS_DATE,
                data_start=9, forn_col=3, max_empty=15, skip_values={"---"},
            )

    meta["Filename"] = filename
    wb.close()

    logger.info(f"[custos_parser] Parsed NFs={len(df_nfs)} rows, Consolidado={len(df_cons)} rows")
    return {"meta": meta, "nfs": df_nfs, "consolidado": df_cons}


def detect_custos_file(file_bytes: bytes, filename: str) -> bool:
    """Detect using zipfile for speed (< 1ms). Imported from custos_template."""
    from .custos_template import detect_custos_file as _detect
    return _detect(file_bytes, filename)

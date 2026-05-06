"""
Parser para Boletim de Medição (BD_Boletim Medição).

Adapted from zip/medicao_parser.py to work with bytes (BytesIO) in addition to file paths.

Estrutura esperada de cada planilha:
- Várias abas MED 01, MED 02, ... (uma por período). Aba 'Auxiliar' é ignorada.
- Cabeçalho (linhas 1-6): obra (D1), fornecedor (G2), período (H6), BM-Nº (N5), vencimento (N6).
- Tabela CONTRATUAL (linha 10 em diante):
    A=item | B=serviço | C=unidade (DIA=efetivo, demais=serviço)
    D=qtde contratual | E=valor unitário contratual | G=valor total contratual
    H=qtde desta medição | I=valor unitário | J=total desta medição
    K=qtde acumulada | L=total acumulado anterior
- Linha de totais da tabela contratual.
- Bloco FATURAMENTO DIRETO com insumos e linha FARKAS CONSTRUTORA.
"""

from __future__ import annotations
import re
from io import BytesIO
from pathlib import Path
from typing import Any
import openpyxl


def _clean(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _round(v: Any, n: int = 2) -> Any:
    if _is_num(v):
        return round(float(v), n)
    return v


def _parse_period(text: Any) -> dict[str, str | None]:
    if not isinstance(text, str):
        return {"raw": None, "inicio": None, "fim": None}
    raw = text.strip()
    m = re.search(r"(\d{2}/\d{2}/\d{4}).*?(\d{2}/\d{2}/\d{4})", raw)
    if m:
        return {"raw": raw, "inicio": m.group(1), "fim": m.group(2)}
    return {"raw": raw, "inicio": None, "fim": None}


def _vencimento_str(v: Any) -> str | None:
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def parse_med_sheet(ws) -> dict | None:
    """Parse one MED sheet. Returns None if empty/invalid."""
    obra = _clean(ws.cell(row=1, column=4).value)          # D1
    fornecedor = _clean(ws.cell(row=2, column=7).value)    # G2
    periodo_raw = ws.cell(row=6, column=8).value            # H6
    bm_numero = ws.cell(row=5, column=14).value             # N5
    vencimento = ws.cell(row=6, column=14).value            # N6

    if not fornecedor or bm_numero is None:
        return None

    periodo = _parse_period(periodo_raw)

    # ---- Tabela contratual (linha 10 em diante) ----
    servicos: list[dict] = []
    row = 10
    last_item_row = row - 1
    while row <= ws.max_row:
        item = ws.cell(row=row, column=1).value
        nome = _clean(ws.cell(row=row, column=2).value)
        if isinstance(item, int) and nome:
            unidade = _clean(ws.cell(row=row, column=3).value)
            tipo = "efetivo" if (isinstance(unidade, str) and unidade.upper() == "DIA") else "servico"
            servicos.append({
                "item": item,
                "nome": nome,
                "unidade": unidade,
                "tipo": tipo,
                "qtde_contratual": _round(ws.cell(row=row, column=4).value, 4),
                "valor_unitario_contratual": _round(ws.cell(row=row, column=5).value),
                "valor_total_contratual": _round(ws.cell(row=row, column=7).value),
                "qtde_medicao": _round(ws.cell(row=row, column=8).value, 4),
                "valor_unitario_medicao": _round(ws.cell(row=row, column=9).value),
                "total_desta_medicao": _round(ws.cell(row=row, column=10).value),
                "qtde_acumulada": _round(ws.cell(row=row, column=11).value, 4),
                "total_acumulado_anterior": _round(ws.cell(row=row, column=12).value),
                "saldo": _round(ws.cell(row=row, column=13).value),
                "percentual": _round(ws.cell(row=row, column=14).value, 4),
            })
            last_item_row = row
        elif item is None and not nome and row > last_item_row:
            empty_streak = 0
            for r2 in range(row, min(row + 4, ws.max_row + 1)):
                if ws.cell(row=r2, column=1).value is None and ws.cell(row=r2, column=2).value is None:
                    empty_streak += 1
            if empty_streak >= 2:
                break
        row += 1

    # ---- Linha de totais da contratual ----
    totais_contratual = {"total_contratual": None, "total_desta_medicao": None, "total_acumulado_atual": None}
    for r in range(last_item_row + 1, last_item_row + 6):
        g = ws.cell(row=r, column=7).value
        j = ws.cell(row=r, column=10).value
        m_label = ws.cell(row=r, column=11).value
        m_val = ws.cell(row=r, column=13).value
        if _is_num(g) and _is_num(j):
            totais_contratual["total_contratual"] = _round(g)
            totais_contratual["total_desta_medicao"] = _round(j)
            if isinstance(m_label, str) and "ACUM" in m_label.upper() and _is_num(m_val):
                totais_contratual["total_acumulado_atual"] = _round(m_val)
            break

    # ---- Bloco FATURAMENTO DIRETO ----
    faturamento = {
        "valor_faturamento_direto": None,
        "retencao_contratual_5pct": None,
        "retencao_impostos": None,
        "valor_fat_direto": None,
        "valor_retencao_tecnica": None,
        "valor_retencao_impostos": None,
        "valor_fat_direto_final": None,
        "valor_retencao_final": None,
        "valor_nf_a_emitir": None,
    }
    insumos: list[dict] = []

    fat_header_row = None
    for r in range(last_item_row + 1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip().upper() == "FATURAMENTO DIRETO":
            fat_header_row = r
            break

    if fat_header_row:
        faturamento["valor_faturamento_direto"] = _round(ws.cell(row=fat_header_row, column=7).value)
        faturamento["retencao_contratual_5pct"] = _round(ws.cell(row=fat_header_row + 1, column=10).value)

        # Itens do faturamento direto
        items_start = fat_header_row + 2
        r = items_start
        while r <= ws.max_row:
            item = ws.cell(row=r, column=1).value
            nome = _clean(ws.cell(row=r, column=2).value)
            qtde = ws.cell(row=r, column=3).value
            valor = ws.cell(row=r, column=5).value
            if isinstance(item, int):
                if nome or _is_num(valor) or _is_num(qtde):
                    insumos.append({
                        "item": item,
                        "fornecedor_produto": nome,
                        "nf_quantidade": qtde,
                        "valor": _round(valor),
                    })
                r += 1
            else:
                break

        # Linha "VALOR FAT. DIRETO"
        for r2 in range(r, min(r + 6, ws.max_row + 1)):
            e = ws.cell(row=r2, column=5).value
            if isinstance(e, str) and "FAT" in e.upper() and "DIRETO" in e.upper():
                faturamento["valor_fat_direto"] = _round(ws.cell(row=r2, column=7).value)
                faturamento["valor_retencao_tecnica"] = _round(ws.cell(row=r2, column=10).value)
                faturamento["valor_retencao_impostos"] = _round(ws.cell(row=r2, column=13).value)
                break

        # Linha "FARKAS CONSTRUTORA" — valor final da NF
        for r2 in range(r, min(r + 8, ws.max_row + 1)):
            b = ws.cell(row=r2, column=2).value
            if isinstance(b, str) and "FARKAS" in b.upper():
                faturamento["valor_fat_direto_final"] = _round(ws.cell(row=r2, column=7).value)
                faturamento["valor_retencao_final"] = _round(ws.cell(row=r2, column=10).value)
                faturamento["valor_nf_a_emitir"] = _round(ws.cell(row=r2, column=13).value)
                break

    return {
        "aba": ws.title,
        "header": {
            "obra": obra,
            "fornecedor": fornecedor,
            "bm_numero": bm_numero,
            "periodo": periodo,
            "vencimento": _vencimento_str(vencimento),
        },
        "servicos": servicos,
        "totais_contratual": totais_contratual,
        "faturamento_direto": faturamento,
        "insumos_faturamento_direto": insumos,
    }


def _parse_wb(wb, filename: str) -> dict:
    medicoes = []
    for sheet_name in wb.sheetnames:
        if sheet_name.upper().startswith("MED "):
            parsed = parse_med_sheet(wb[sheet_name])
            if parsed:
                medicoes.append(parsed)

    fornecedor = medicoes[0]["header"]["fornecedor"] if medicoes else None
    obra = medicoes[0]["header"]["obra"] if medicoes else None
    return {
        "arquivo": filename,
        "fornecedor": fornecedor,
        "obra": obra,
        "qtd_medicoes": len(medicoes),
        "medicoes": medicoes,
    }


def parse_workbook_from_bytes(content: bytes, filename: str) -> dict:
    """Parse Boletim de Medição from raw bytes (used in upload handler)."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    return _parse_wb(wb, filename)


def parse_workbook(path: str | Path) -> dict:
    """Parse Boletim de Medição from file path."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    return _parse_wb(wb, path.name)


def detect_medicao_file(file_bytes: bytes, filename: str) -> bool:
    """
    Detect if a file is a Boletim de Medição spreadsheet:
    1. Filename contains "boletim" + "medi" (case-insensitive, handles "Medição"/"Medicao")
    2. Fallback: workbook has at least one MED sheet with fornecedor at G2 and BM-Nº at N5
    """
    name_lower = filename.lower()
    if "boletim" in name_lower and "medi" in name_lower:
        return True

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        med_sheets = [s for s in wb.sheetnames if s.upper().startswith("MED ")]
        if med_sheets:
            ws = wb[med_sheets[0]]
            fornecedor = ws.cell(row=2, column=7).value
            bm = ws.cell(row=5, column=14).value
            wb.close()
            if fornecedor and bm is not None:
                return True
        else:
            wb.close()
    except Exception:
        pass
    return False


if __name__ == "__main__":
    import sys
    import json
    out = parse_workbook(sys.argv[1])
    print(json.dumps(out, indent=2, ensure_ascii=False, default=str))

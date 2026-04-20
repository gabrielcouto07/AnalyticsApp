"""
Efetivo Template — Template definition for "Controle de Efetivo" (Workforce Control) data.

This template handles construction site workforce tracking spreadsheets with:
- Multiple Fornecedores (contractors/suppliers)
- Worker functions (Pedreiro, Servente, Armador, etc.)
- Daily worker counts per function
- Monthly totals (Diárias)

The Excel layout is NON-STANDARD — one sheet per month, with:
  Row 1: Obra name + "Controle de Efetivo"
  Row 2: "Mês:" + month name (col AH/AI)
  Row 3: "Função:" header + day numbers 1-31 + "Diárias Totais:"
  Row 6+: Fornecedor blocks (name row, then serviço rows, separated by empty rows)
  Year from filename (e.g., "Efetivo_2026.xlsx")
"""

from typing import Dict, List, Any

# ============================================================================
# EFETIVO TEMPLATE DEFINITION
# ============================================================================

EFETIVO_TEMPLATE = {
    "name": "Efetivo - Controle de Mão de Obra",
    "description": "Rastreamento profissional de mão de obra em canteiro. Análise de diárias por fornecedor, função e período com tendências temporais.",
    "icon": "👷",
    "color": "#f5a623",
    "tags": ["mão-de-obra", "canteiro", "fornecedores", "temporal"],
    "key_metrics": [
        {
            "name": "Total Diárias",
            "field": "Quantidade",
            "description": "Soma de todas as diárias trabalhadas",
            "type": "sum",
            "icon": "📊",
        },
        {
            "name": "Fornecedores Ativos",
            "field": "Fornecedor",
            "description": "Quantidade de fornecedores únicos",
            "type": "unique_count",
            "icon": "🤝",
        },
        {
            "name": "Funções",
            "field": "Funcao",
            "description": "Tipos de funções executadas",
            "type": "unique_count",
            "icon": "🛠️",
        },
        {
            "name": "Média de Efetivo",
            "field": "Quantidade",
            "description": "Média diária de trabalhadores",
            "type": "average",
            "icon": "📈",
        },
        {
            "name": "Pico de Efetivo",
            "field": "Quantidade",
            "description": "Dia com maior concentração de diárias",
            "type": "max",
            "icon": "⚡",
        },
    ],
    "required_columns": [
        "Obra",
        "Ano",
        "Mes",
        "MesNome",
        "Fornecedor",
        "Funcao",
        "Dia",
        "Quantidade",
        "DiariasTotal",
        "Data",
    ],
    "visualizations": [
        {
            "type": "bar",
            "title": "Diárias por Fornecedor",
            "description": "Distribuição total de trabalho entre fornecedores",
            "field": "Fornecedor",
            "value_field": "Quantidade",
            "icon": "📊",
        },
        {
            "type": "pie",
            "title": "Participação por Função",
            "description": "Percentual de cada especialidade",
            "field": "Funcao",
            "value_field": "Quantidade",
            "icon": "🥧",
        },
        {
            "type": "line",
            "title": "Evolução Temporal de Efetivo",
            "description": "Tendência de mão de obra ao longo do período",
            "date_field": "Data",
            "value_field": "Quantidade",
            "icon": "📈",
        },
        {
            "type": "bar",
            "title": "Diárias por Mês",
            "description": "Comparação mensal de volume de trabalho",
            "field": "MesNome",
            "value_field": "Quantidade",
            "icon": "📅",
        },
        {
            "type": "heatmap",
            "title": "Mapa de Calor: Fornecedor vs Função",
            "description": "Análise da composição de cada fornecedor",
            "x_field": "Fornecedor",
            "y_field": "Funcao",
            "value_field": "Quantidade",
            "icon": "🔥",
        },
    ],
    "filters": [
        {"field": "Fornecedor", "type": "multi_select"},
        {"field": "Funcao", "type": "multi_select"},
        {"field": "MesNome", "type": "multi_select"},
        {"field": "Obra", "type": "single_select"},
        {"field": "Data", "type": "date_range"},
    ],
    "sample_columns": {
        "Obra": "text",
        "Ano": "numeric",
        "Mes": "numeric",
        "MesNome": "text",
        "Fornecedor": "text",
        "Funcao": "text",
        "Dia": "numeric",
        "Quantidade": "numeric",
        "DiariasTotal": "numeric",
        "Data": "date",
    },
    # Custom flag: tells the upload router to use the efetivo parser
    "custom_parser": "efetivo",
}


# ============================================================================
# COLUMN MAPPING — How to detect this is an Efetivo file
# ============================================================================

EFETIVO_DETECTION_KEYWORDS = {
    "strong": [
        "controle de efetivo", "função:", "diárias totais",
        "pedreiro", "servente", "encarregado", "armador",
        "meio oficial", "betoneiro", "guincheiro",
    ],
    "medium": [
        "obra_", "fornecedor", "mês:", "carpinteiro",
        "eletricista", "bombeiro", "poceiro",
    ],
    "weak": [
        "efetivo", "diárias", "canteiro",
    ],
}


def detect_efetivo_file(file_bytes: bytes, filename: str) -> bool:
    """
    Detect if a file is an Efetivo spreadsheet by checking:
    1. Filename contains 'efetivo' (case-insensitive)
    2. Cell A1 contains 'Obra_' pattern
    3. Cell B1 contains 'Controle de Efetivo'
    4. Row 3 col A contains 'Função:'
    """
    name_lower = filename.lower()
    if "efetivo" in name_lower:
        return True

    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        a1 = str(ws.cell(row=1, column=1).value or "").lower()
        b1 = str(ws.cell(row=1, column=2).value or "").lower()
        a3 = str(ws.cell(row=3, column=1).value or "").lower()

        wb.close()

        if "obra" in a1 and "controle" in b1 and "efetivo" in b1:
            return True
        if "função" in a3 or "funcao" in a3:
            return True
    except Exception:
        pass

    return False


def get_efetivo_template() -> Dict[str, Any]:
    """Return the Efetivo template definition."""
    return EFETIVO_TEMPLATE


def get_efetivo_suggestions_score(columns: List[str]) -> float:
    """
    Score how well a set of column names matches the Efetivo template.
    Used by the template suggestion engine.
    """
    col_lower = [c.lower().strip() for c in columns]
    score = 0.0

    for col in col_lower:
        for kw in EFETIVO_DETECTION_KEYWORDS["strong"]:
            if kw in col or col in kw:
                score += 3.0
        for kw in EFETIVO_DETECTION_KEYWORDS["medium"]:
            if kw in col or col in kw:
                score += 1.5
        for kw in EFETIVO_DETECTION_KEYWORDS["weak"]:
            if kw in col or col in kw:
                score += 1.0

    return score
